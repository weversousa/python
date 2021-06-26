from openpyxl import load_workbook

arquivo_excel = load_workbook('excel/arquivo_criado.xlsx')

arquivo_excel.create_sheet('Estados')

print(arquivo_excel.sheetnames)

arquivo_excel.close()
