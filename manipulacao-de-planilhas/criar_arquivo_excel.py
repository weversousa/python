from openpyxl import Workbook

arquivo_excel = Workbook()

# Ao ser instanciada por padrão é criado a planilha1.
planilha_1 = arquivo_excel.active
# -----------------------------------------------------------------------------

# E você pode alterar o nome dela.
planilha_1.title = 'Produtos'
# -----------------------------------------------------------------------------

# Agora podemos tratar como um Excel, e atribuir valores as células.
planilha_1['A1'] = 'Nome'
planilha_1['B1'] = 'Qtde'
planilha_1['C1'] = 'Preço'
# -----------------------------------------------------------------------------

# Atribuir valores através de Tuplas. Cada Tupla será uma linha da planilha.
# Para isso usamos o método append() .
produtos = [
    ('Arroz', 1, 28.30),
    ('Feijão', 1, 9.70),
    ('Café', 1, 11.30)
]

for produto in produtos:
    planilha_1.append(produto)
# -----------------------------------------------------------------------------

# O método cell() é outra forma de adicionar valores.
# Para isso devemos passar a Coluna, Linha e o Valor como parâmetros.
planilha_1.cell(row=5, column=1, value='Requeijão')
planilha_1.cell(row=5, column=2, value=1)
planilha_1.cell(row=5, column=3, value=5.40)
# -----------------------------------------------------------------------------

# Inserindo fórmulas.
planilha_1['C6'] = '=SOMA(C2:C5)'
# -----------------------------------------------------------------------------

# a2 receberá um objeto e para acessar o seu valor usamos o atributo value.
a2 = planilha_1['A2']
print(a2.value)

# Resultado:
# Arroz
# -----------------------------------------------------------------------------

# O método cell() também retorna o valor de tal célula.
a3 = planilha_1.cell(column=1, row=3)
print(a3.value)

# Resultado:
# Feijão
# -----------------------------------------------------------------------------

# Para percorrer uma planilha inteira é preciso saber quantas linhas e colunas
# ela tem. Para isso é usado os atributos max_row e max_column.
# Após isso podemos fazer um for do início ao fim da planilha.
total_de_linhas_usadas = planilha_1.max_row
total_de_colunas_usadas = planilha_1.max_column

for linha in range(1, total_de_linhas_usadas, 1):
    for coluna in range(1, total_de_colunas_usadas, 1):
        print(planilha_1.cell(row=linha, column=coluna).value)

# Resultado:
# Nome
# Qtde
# Arroz
# 1
# Feijão
# 1
# Café
# 1
# Requeijão
# 1
# -----------------------------------------------------------------------------

# Trabalhando com mais de uma planilha

arquivo_excel.create_sheet('Planilha2')
arquivo_excel.create_sheet('Planilha3')

planilha2 = arquivo_excel['Planilha2']
planilha3 = arquivo_excel['Planilha3']

planilha2.append(('Nome', 'Sobrenome'))
planilha2.append(('Weverton', 'Teixeira'))

planilha3.append(('Cidade', 'Estado'))
planilha3.append(('Osasco', 'SP'))
# -----------------------------------------------------------------------------

# Quando criamos uma classe Worlbook(), estamos criando uma área de trabalho
# na memória.
# O arquivo Excel somente é criado quando usamos a função save()
arquivo_excel.save('arquivo_criado.xlsx')
arquivo_excel.close()
