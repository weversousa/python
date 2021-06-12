from openpyxl import Workbook

arquivo_excel = Workbook()

planilha_1 = arquivo_excel.active
# -----------------------------------------------------------------------------

''' Por padrão é criado planilha1, e você pode alterar o nome '''

planilha_1.title = 'Produtos'
# -----------------------------------------------------------------------------

planilha_1['A1'] = 'Nome'
planilha_1['B1'] = 'Qtde'
planilha_1['C1'] = 'Preço'
# -----------------------------------------------------------------------------

produtos = [
    ('Arroz', 1, 28.30),
    ('Feijão', 1, 9.70),
    ('Café', 1, 11.30)
]

for produto in produtos:
    planilha_1.append(produto)
# -----------------------------------------------------------------------------

planilha_1.cell(row=5, column=1, value='Requeijão')
planilha_1.cell(row=5, column=2, value=1)
planilha_1.cell(row=5, column=3, value=5.40)
# -----------------------------------------------------------------------------

planilha_1['C6'] = '=SOMA(C2:C5)'
# -----------------------------------------------------------------------------

a2 = planilha_1['A2']
print(a2.value)

# Resultado:
# Arroz
# -----------------------------------------------------------------------------

a3 = planilha_1.cell(column=1, row=3)
print(a3.value)

# Resultado:
# Feijão
# -----------------------------------------------------------------------------

total_de_linhas_usadas = planilha_1.max_row
total_de_colunas_usadas = planilha_1.max_column

for linha in range(1, total_de_linhas_usadas, 1):
    for coluna in range(1, total_de_colunas_usadas, 1):
        print(planilha_1.cell(row=linha, column=coluna).value)

# Resultado:
# Nome
# Qtde
# Arroz
# 1
# Feijão
# 1
# Café
# 1
# Requeijão
# 1
# -----------------------------------------------------------------------------

arquivo_excel.save('excel/arquivo_criado.xlsx')
arquivo_excel.close()
